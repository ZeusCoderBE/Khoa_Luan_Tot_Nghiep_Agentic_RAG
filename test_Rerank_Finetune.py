import pandas as pd
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Đọc dữ liệu từ file CSV
csv_path = 'source_research/Eval_System/data/data_processed/giaoduc_merge.csv'
df = pd.read_csv(csv_path)
df = df.head(150)

# Chuẩn bị file Excel
excel_path = 'test_Rerank_Finetune.xlsx'
wb = Workbook()
ws = wb.active
ws.title = 'Results'
ws.append(['question', 'answer', 'answerSystemFinetune'])

# Đảm bảo font tiếng Việt không bị lỗi
for col in range(1, 4):
    ws.column_dimensions[get_column_letter(col)].width = 50

# Gửi request và lưu kết quả
url = 'http://127.0.0.1:8000/api/chat/chatbot-with-gemini'

for idx, row in df.iterrows():
    question = str(row['question'])
    answer = str(row['answer'])
    try:
        response = requests.post(url, json={'query': question})
        if response.status_code == 200:
            data = response.json()
            answerSystemFinetune = data.get('answer', '')
        else:
            answerSystemFinetune = f"Lỗi API: {response.status_code}"
    except Exception as e:
        answerSystemFinetune = f"Lỗi: {e}"
    ws.append([question, answer, answerSystemFinetune])
    wb.save(excel_path)
    print(f"Đã thực hiện xong câu: {idx+1}") 